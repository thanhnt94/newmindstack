"""Helper utilities for working with Excel files in content management flows."""
from __future__ import annotations

from typing import List, Tuple, Optional, Any
import logging

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def read_excel_with_formulas(
    file_path: str,
    sheet_name: str = 'Data',
    **pandas_kwargs
) -> pd.DataFrame:
    """
    Read an Excel file and automatically evaluate formulas.
    
    This function attempts to get computed values from formulas instead of
    the formula strings themselves. It uses openpyxl's data_only mode first
    (which reads cached values if the file was saved by Excel), then falls
    back to the formulas library for actual computation if needed.
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to read (default: 'Data')
        **pandas_kwargs: Additional arguments to pass to pd.read_excel
        
    Returns:
        pd.DataFrame: DataFrame with computed values
    """
    try:
        # Strategy 1: Use openpyxl with data_only=True
        # This reads cached formula results (if file was saved by Excel)
        wb = load_workbook(file_path, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")
        
        ws = wb[sheet_name]
        
        # Convert worksheet to list of lists
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        
        if not data:
            return pd.DataFrame()
        
        # First row is header - filter out None values and convert to str
        raw_headers = data[0]
        headers = []
        for i, h in enumerate(raw_headers):
            if h is None:
                # Skip None columns (empty header cells)
                continue
            headers.append(str(h).strip() if h else f"Column_{i}")
        
        # Filter rows to only include columns with valid headers
        valid_indices = [i for i, h in enumerate(raw_headers) if h is not None]
        rows = []
        for row in data[1:]:
            filtered_row = [row[i] if i < len(row) else None for i in valid_indices]
            rows.append(filtered_row)
        
        # Check if any cells have None values that might be uncalculated formulas
        # by also loading without data_only and comparing
        wb_formulas = load_workbook(file_path, data_only=False)
        ws_formulas = wb_formulas[sheet_name]
        
        has_uncalculated = False
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=0):
            for col_idx, cell_value in enumerate(row):
                if cell_value is None:
                    # Check if the original cell has a formula
                    formula_cell = ws_formulas.cell(row=row_idx + 2, column=col_idx + 1)
                    if formula_cell.value and str(formula_cell.value).startswith('='):
                        has_uncalculated = True
                        break
            if has_uncalculated:
                break
        
        wb.close()
        wb_formulas.close()
        
        # If we found uncalculated formulas, try using formulas library
        if has_uncalculated:
            try:
                import formulas
                logger.info(f"Using formulas library to compute Excel formulas for {file_path}")
                
                xl_model = formulas.ExcelModel().loads(file_path).finish()
                xl_model.calculate()
                
                # Re-read the computed values
                computed_data = []
                for row_idx in range(len(rows)):
                    computed_row = []
                    for col_idx in range(len(headers)):
                        # formulas uses 1-based indexing for rows (skip header)
                        cell_ref = f"'{sheet_name}'!{_get_column_letter(col_idx + 1)}{row_idx + 2}"
                        try:
                            value = xl_model.books[file_path][cell_ref].value
                            # Handle numpy types
                            if hasattr(value, 'item'):
                                value = value.item()
                            computed_row.append(value)
                        except (KeyError, AttributeError):
                            # Fallback to original value
                            computed_row.append(rows[row_idx][col_idx] if col_idx < len(rows[row_idx]) else None)
                    computed_data.append(computed_row)
                
                return pd.DataFrame(computed_data, columns=headers)
                
            except ImportError:
                logger.warning(
                    "formulas library not installed. Some Excel formulas may not be computed. "
                    "Install with: pip install formulas"
                )
            except Exception as e:
                logger.warning(f"Failed to compute formulas with formulas library: {e}")
        
        # Return DataFrame from cached values
        return pd.DataFrame(rows, columns=headers)
        
    except Exception as e:
        logger.warning(f"Error reading Excel with formula evaluation: {e}. Falling back to pandas.")
        # Fallback to standard pandas read
        return pd.read_excel(file_path, sheet_name=sheet_name, **pandas_kwargs)


def _get_column_letter(col_idx: int) -> str:
    """Convert 1-based column index to Excel column letter (A, B, ..., Z, AA, AB, ...)."""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def extract_info_sheet_mapping(file_path: str) -> Tuple[dict[str, str], List[str]]:
    """Return mapping data from the ``Info`` sheet and any warnings encountered.

    The function tries to be forgiving with partially filled sheets. It accepts
    missing values, trims whitespace around keys/values and records any
    recoverable issues as human readable warnings so that callers can present
    actionable feedback to end users.
    """
    warnings: List[str] = []
    try:
        df_info = read_excel_with_formulas(file_path, sheet_name="Info")
    except ValueError as ve:
        # read_excel_with_formulas raises ValueError if sheet not found
        if "not found" in str(ve).lower():
            warnings.append(
                "Không tìm thấy sheet 'Info'. Hãy giữ nguyên sheet này khi tải file từ hệ thống."
            )
        else:
            warnings.append(f"Không thể đọc sheet 'Info': {ve}")
        return {}, warnings
    except Exception as exc:  # pragma: no cover - defensive branch
        warnings.append(f"Không thể đọc sheet 'Info': {exc}")
        return {}, warnings

    if df_info.empty:
        warnings.append("Sheet 'Info' đang trống. Điền dữ liệu vào hai cột 'Key' và 'Value'.")
        return {}, warnings

    normalized_columns = {
        str(column).strip().lower(): column
        for column in df_info.columns
        if isinstance(column, str)
    }
    key_column_name = normalized_columns.get("key")
    value_column_name = normalized_columns.get("value")

    if not key_column_name:
        warnings.append("Sheet 'Info' thiếu cột 'Key'. Đảm bảo cột đầu tiên có tiêu đề 'Key'.")
        return {}, warnings

    if not value_column_name:
        remaining_columns = [column for column in df_info.columns if column != key_column_name]
        if remaining_columns:
            value_column_name = remaining_columns[0]
            warnings.append(
                "Sheet 'Info' thiếu cột 'Value'. Đang dùng tạm cột "
                f"'{value_column_name}' và cần đổi tên thành 'Value' để tránh lỗi."
            )
        else:
            warnings.append("Sheet 'Info' thiếu cột 'Value'. Hãy thêm cột thứ hai với tiêu đề 'Value'.")
            return {}, warnings

    info_mapping: dict[str, str] = {}
    for _, row in df_info.iterrows():
        raw_key = row.get(key_column_name)
        if pd.isna(raw_key):
            continue
        key = str(raw_key).strip()
        if not key:
            continue

        raw_value = row.get(value_column_name)
        if pd.isna(raw_value):
            info_mapping[key] = ""
        else:
            info_mapping[key] = str(raw_value).strip()

    if not info_mapping:
        warnings.append(
            "Sheet 'Info' không có dòng hợp lệ. Kiểm tra lại cột 'Key' và 'Value'."
        )

    return info_mapping, warnings


def format_info_warnings(warnings: List[str]) -> str:
    """Turn a list of warning strings into a short, human friendly sentence."""
    if not warnings:
        return ""
    if len(warnings) == 1:
        return warnings[0]
    return " ".join(f"{index + 1}. {message}" for index, message in enumerate(warnings))


# --- Content Parsing Utilities (Unified) ---

# Action normalization aliases
ACTION_ALIASES = {
    'delete': {'delete', 'remove'},
    'skip': {'skip', 'keep', 'none', 'ignore', 'nochange', 'unchanged', 
             'giu nguyen', 'giu-nguyen', 'giu_nguyen'},
    'create': {'create', 'new', 'add', 'insert'},
    'update': {'update', 'upsert', 'edit', 'modify'},
}


def classify_columns(
    columns: List[str],
    standard_columns: set[str],
    system_columns: set[str],
    ai_columns: set[str]
) -> dict[str, List[str]]:
    """Classify DataFrame columns into categories."""
    columns_set = set(columns)
    all_known = standard_columns | system_columns | ai_columns
    
    return {
        'standard': sorted([c for c in columns if c in standard_columns]),
        'system': sorted([c for c in columns if c in system_columns]),
        'ai': sorted([c for c in columns if c in ai_columns]),
        'custom': sorted([c for c in columns if c not in all_known]),
        'missing_required': [c for c in ['front', 'back'] if c not in columns_set],
        'all': sorted(list(columns))
    }


def normalize_action(
    raw_action: Optional[str],
    has_item_id: bool
) -> str:
    """Normalize action string to standard values: 'create', 'update', 'delete', or 'skip'."""
    value = (raw_action or '').strip().lower()
    
    if value:
        for normalized, alias_values in ACTION_ALIASES.items():
            if value in alias_values:
                if normalized == 'create' and has_item_id:
                    return 'update'
                if normalized == 'update' and not has_item_id:
                    return 'create'
                return normalized
    
    return 'update' if has_item_id else 'create'


def get_cell_value(
    row_data: Any,
    column_name: str,
    columns: List[str]
) -> Optional[str]:
    """Safely get a cell value from row data (dict or Series)."""
    if column_name not in columns:
        return None
    
    # Handle dict or pandas Series
    if isinstance(row_data, dict):
        value = row_data.get(column_name)
    else:
        value = row_data[column_name]
    
    if pd.isna(value):
        return None
    
    return str(value).strip()


def parse_column_pairs(pairs_string: str) -> List[dict[str, str]]:
    """Parse column pair configuration string "q:a | q2:a2"."""
    if not pairs_string:
        return []
    
    pairs_list = []
    raw_pairs = str(pairs_string).split('|')
    
    for raw_pair in raw_pairs:
        parts = raw_pair.split(':')
        if len(parts) == 2:
            q_col = parts[0].strip()
            a_col = parts[1].strip()
            if q_col and a_col:
                pairs_list.append({'q': q_col, 'a': a_col})
    
    return pairs_list
